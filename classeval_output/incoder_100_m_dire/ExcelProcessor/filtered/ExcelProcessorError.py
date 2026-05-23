import openpyxl
class ExcelProcessor: 
    def __init__(self):
        pass



    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.get_sheet_by_name('Sheet1')
        data = []
        for row in sheet.rows:
            data.append([cell.value for cell in row])
        return data


    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        writer = pd.ExcelWriter(file_name)
        writer.book = load_workbook(file_name)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        writer.sheets['Sheet1'].write(0,0,'Name')
        writer.sheets['Sheet1'].write(0,1,'Age')
        writer.sheets['Sheet1'].write(0,2,'Country')
        writer.sheets['Sheet1'].write(0,3,'Age')
        writer.sheets['Sheet1'].write(0,4,'Country')
        writer.sheets['Sheet1'].write(0,5,'Age')
        writer.sheets['Sheet1'].write(0,6,'Country')
        writer.sheets['Sheet1'].write(0,7,'Age')
        writer.sheets['Sheet1'].write(0,8,'Country')
        writer.sheets['Sheet1'].write(0,9,'Age')
        writer.sheets['Sheet1'].write(0,10,'Country')
        writer.sheets['Sheet1'].write(0,11,'Age')
        writer.sheets['Sheet1'].write(0,12,'Country')
        writer.sheets['Sheet1'].write(0,13,'Age')
        writer.sheets['Sheet1'].write(0,14,'Country')
        writer.sheets['Sheet1'].write(0,15,'Age')
        writer.sheets['Sheet1'].write(0,16,'Country')
        writer.sheets['Sheet1'].write(0,17,'Age')
        writer.sheets['Sheet1'].write(0,18,'Country')
        writer.sheets['Sheet1'].write(0,19,'Age')
        writer.sheets['Sheet1'].write(0,20,'Country')
        writer.sheets['Sheet1'].write(0,21,'Age')
        writer.sheets['Sheet1'].write(0,22,'Country')
        writer.sheets['Sheet1'].write(0,23,'Age')
        writer.sheets['Sheet1'].write(0,24,'Country')
        writer.sheets['Sheet1'].write(0,25,'Age')
        writer.sheets['Sheet1'].write(0,26,'Country')
        writer.sheets['Sheet1'].write(0,27,'Age')
        writer.sheets['Sheet1'].write(0,28,'Country')
        writer.sheets['Sheet1'].write(0,29,'Age')
        writer.sheets['Sheet1'].write(0,30,'Country')
        writer.sheets['Sheet1'].write(0,31,'Age')
        writer.sheets['Sheet1'].write(0,32,'Country')
        writer.sheets['Sheet1'].write(0,33,'Age')
        writer.sheets['Sheet1'].write(0,34,'Country')
        writer.sheets['Sheet1'].write(0,35,'Age')
        writer.sheets['Sheet1'].write(0,36,'Country')
        writer.sheets['Sheet1'].write(0,37,'Age')
        writer.sheets['Sheet1'].write(0,38,'Country')
        writer.sheets['Sheet1'].write(0,39,'Age')
        writer.sheets['Sheet1'].write(0,40,'Country')
        writer.sheets['Sheet1'].write(0,41,'Age')
        writer.sheets['Sheet1'].write(0,42,'Country')
        writer.sheets['Sheet1'].write(0,43,'Age')
        writer.sheets['Sheet1'].write(0,44,'Country')
        writer.sheets['Sheet1'].write(0,45,'Age')
        writer.sheets['Sheet1'].write(0,46,'Country')
        writer.sheets['Sheet1'].write(0,47,'Age')
        writer.sheets['Sheet1'].write(0,48,'Country')
        writer.sheets['Sheet1'].write(0,49,'Age')
        writer.sheets['Sheet1'].write(0,50,'Country')
        writer.sheets['Sheet1'].write(0,51,'Age')
        writer.sheets['Sheet1'].write(0,52,'Country')
        writer.sheets['Sheet1'].write(0,53,'Age')
        writer.sheets['Sheet1'].write(0,54,'Country')
        writer.sheets['Sheet1'].write(0,55,'Age')
        writer.sheets['Sheet1'].write(0,56,'Country')
        writer.sheets['Sheet1'].write(0,57,'Age')
        writer.sheets['Sheet1'].write(0,58,'Country')
        writer.sheets['Sheet1'].write(0,59,'Age')
        writer.sheets['Sheet1'].write(0,60,'Country')
        writer.sheets['Sheet1'].write(0,61,'Age')
        writer.sheets['Sheet1'].write(0,62,'Country')
        writer.sheets['Sheet1'].write(0,63,'Age')
        writer.sheets['Sheet1'].write(0,64,'Country')
        writer.sheets['Sheet1'].write(0,65,'Age')
        writer.sheets['Sheet1'].write(0,66,'Country')
        writer.sheets['Sheet1'].write(0,67,'Age')
        writer.sheets['Sheet1'].write(0,68,'Country')
        writer.sheets['Sheet1'].write(0,69,'Age')
        writer.sheets['Sheet1'].write(0,70,'Country')
        writer.sheets['Sheet1'].write(0,71,'Age')
        writer.sheets['Sheet1'].write(0,72,'Country')
        writer.sheets['Sheet1'].write(0,73,'Age')
        writer.sheets['Sheet1'].write(0,74,'Country')
        writer.sheets['Sheet1'].write(0,75,'Age')
        writer.sheets['Sheet1'].write(0,76,'Country')
        writer.sheets['Sheet1'].write(0,77,'Age')
        writer.sheets['Sheet1'].write(0,78,'Country')
        writer.sheets['Sheet1'].write(0,79,'Age')
        writer.sheets['Sheet1'].write(0,80,'Country')
        writer.sheets['Sheet1'].write(0,81,'Age')
        writer.sheets['Sheet1'].write(0,82,'Country')
        writer.sheets['Sheet1'].write(0,83,'Age')
        writer.sheets['Sheet1'].write(0,84,'Country')
        writer.sheets['Sheet1'].write(0,85,'Age')
        writer.sheets['Sheet1'].write(0,86,'Country')
        writer.sheets['Sheet1'].write(0,87,'Age')
        writer.sheets['Sheet1'].write(0,88,'Country')
        writer.sheets['Sheet1'].write(0,89,'Age')
        writer.sheets['Sheet1'].write(0,90,'Country')
        writer.sheets['Sheet1'].write(0,91,'Age')
        writer.sheets['Sheet1'].write(0,92,'Country')
        writer.sheets['Sheet1'].write(0,93,'Age')
        writer.sheets['Sheet1'].write(0,94,'Country')
        writer.sheets['Sheet1'].write(0,95,'Age')
        writer.sheets['Sheet1'].write(0,96,'Country')
        writer.sheets['Sheet1'].write(0,97,'Age')
        writer.sheets['Sheet1'].write(0,98,'Country')
        writer.sheets['Sheet1'].write(0,99,'Age')
        writer.sheets['Sheet1'].write(0,100,'Country')
        writer.sheets['Sheet1'].write(0,101,'Age')
        writer.sheets['Sheet1'].write(0,102,'Country')
        writer.sheets['Sheet1'].write(0,103,'Age')
        writer.sheets['Sheet1'].write(0,104,'Country')
        writer.sheets['Sheet1'].write(0,105,'Age')
        writer.sheets['Sheet1'].write(0,106,'Country')
        writer.sheets['Sheet1'].write(0,107,'Age')
        writer.sheets['Sheet1'].write(0,108,"Country")
        writer.sheets['Sheet1'].write(0,109,'Age')
        writer.sheets['Sheet1'].write(0,110,'Country')
        writer.sheets['Sheet1'].write(0,111,'Age')
        writer.sheets['Sheet1'].write(0,112,'Country')
        writer.sheets['Sheet1'].write(0,113,'Age')
        writer.sheets['Sheet1'].write(0,114,'Country')
        writer.sheets['Sheet1'].write(0,115,'Age')
        writer.sheets['Sheet1'].write(0,116,'Country')
        writer.sheets['Sheet1'].write(0,117,'Age')
        writer.sheets['Sheet1'].write(0,118,'Country')
        writer.sheets['Sheet1'].write(0,119,'Age')
        writer.sheets['Sheet1'].write(0,120,'Country')
        writer.sheets['Sheet1'].write(0,121,'Age')
        writer.sheets['Sheet1'].write(0,122,'Country')
        writer.sheets['Sheet1'].write(0,123,'Age')
        writer.sheets['Sheet1'].write(0,124,'Country')
        writer.sheets['Sheet1'].write(0,125,'Age')
        writer.sheets['Sheet1'].write(0,126,'Country')
        writer.sheets['Sheet1'].write(0,127,'Age')
        writer.sheets['Sheet1'].write(0,


    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        workbook = openpyxl.load_workbook(save_file_name)
        worksheet = workbook.active
        worksheet.column_dimensions[N].width = 256
        worksheet.column_dimensions[N].auto_size = True
        worksheet.column_dimensions[N].number_format = '@'
        worksheet.column_dimensions[N].number_format_str = '@'
        worksheet.column_dimensions[N].alignment = openpyxl.styles.Alignment()
        worksheet.column_dimensions[N].alignment.horizontal = 'center'
        worksheet.column_dimensions[N].alignment.vertical = 'center'
        worksheet.column_dimensions[N].alignment.wrap_text = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_cells = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_pages = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_width = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_height = True
        worksheet.column_dimensions[N].alignment.shrink_to_fit_content = True
        workbook.save(save_file_name)