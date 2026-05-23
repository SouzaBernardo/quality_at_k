Please complete the class ExcelProcessor in the following code.

```python
import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        try:
            workbook.save(file_name)
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        data = self.read_excel(save_file_name)
        processed_data = []
        for row in data:
            processed_row = []
            for i, value in enumerate(row):
                if i == N - 1:
                    processed_row.append(str(value).upper())
                else:
                    processed_row.append(value)
            processed_data.append(tuple(processed_row))
        return self.write_excel(processed_data, save_file_name)
```

This completes the class ExcelProcessor with the required methods: `read_excel`, `write_excel`, and `process_excel_data`. The `read_excel` method reads data from an Excel file and returns it as a list. The `write_excel` method writes data to a specified Excel file. The `process_excel_data` method changes the specified column in the Excel file to uppercase and saves the processed data to a new file.