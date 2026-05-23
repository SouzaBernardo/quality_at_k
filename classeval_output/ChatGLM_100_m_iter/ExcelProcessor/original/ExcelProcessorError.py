import openpyxl
class ExcelProcessor:  
    """
    This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass



    def read_excel(self, file_name):
        if self.file_path is None:
            self.file_path = os.path.join(os.path.dirname(__file__), file_name)
        
        with openpyxl.load_workbook(self.file_path) as workbook:
            data = []
            for row in workbook.rows:
                for cell in row:
                    data.append(cell.value)
            return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        if not data:
            return 0
        
        file_path = os.path.join(os.path.dirname(__file__), file_name)
        workbook = openpyxl.load_workbook(file_path)
        
        for row in data:
            for cell in row:
                cell_value = cell.value
                if cell_value not in workbook. cells:
                    workbook. cells[cell_value] = []
                workbook. cells[cell_value].append(cell_value)
        
        workbook. save(file_path)
        return 1

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        if not N:
            return 0, save_file_name
        
        file_path = os.path.join(os.path.dirname(__file__), save_file_name)
        workbook = openpyxl.load_workbook(file_path)
        
        for row in data:
            for cell in row:
                if cell.value not in workbook. cells:
                    workbook. cells[cell.value] = []
                workbook. cells[cell.value].append(cell.value)
        
        for row in workbook.rows[N:]:
            for cell in row:
                cell_value = cell.value
                if cell_value not in workbook. cells:
                    workbook. cells[cell_value] = []
                workbook. cells[cell_value].append(cell_value.upper())
        
        workbook. save(file_path)
        return 1, save_file_name