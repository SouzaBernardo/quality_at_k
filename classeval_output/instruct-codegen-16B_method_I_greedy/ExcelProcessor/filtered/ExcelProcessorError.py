import openpyxl
class ExcelProcessor:  
    """
    This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass



    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        # open the excel file
        wb = openpyxl.load_workbook(file_name)
        # get the active sheet
        sheet = wb.active
        # read the data
        data = []
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        # close the file
        wb.close()
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        # open the excel file
        wb = openpyxl.Workbook()
        # get the active sheet
        sheet = wb.active
        # write the data
        for row in data:
            row_data = []
            for cell in row:
                row_data.append(cell)
            sheet.append(row_data)
        # save the file
        wb.save(file_name)
        return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        # open the source file
        wb = openpyxl.load_workbook(save_file_name)
        # get the active sheet
        sheet = wb.active
        # get the column to be changed
        column = sheet.cell(row=1, column=N).value
        # change the column to uppercase
        column = column.upper()
        # write the changed column back to the file
        sheet.cell(row=1, column=N).value = column
        # save the file
        wb.save(save_file_name)
        return (0, save_file_name)