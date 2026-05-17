import openpyxl
class ExcelProcessor: 
    def __init__(self):
        pass


    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        # Load workbook
        workbook = openpyxl.load_workbook(file_name)
    
        # Select the active worksheet
        sheet = workbook.active
    
        # Initialize data list
        data = []
    
        # Iterate over the rows and add each row's data to the data list
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
    
        # Return the data
        return data
    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            # Create a workbook and select the active worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
    
            # Write data to the worksheet
            for row in data:
                worksheet.append(row)
    
            # Save the workbook
            workbook.save(file_name)
    
            # Return 1 to indicate successful writing
            return 1
        except Exception as e:
            print(f"An error occurred: {e}")
            # Return 0 to indicate failed writing
            return 0
    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        data = self.read_excel(save_file_name)
        for row in data:
            row[N-1] = str(row[N-1]).upper() if row[N-1] else row[N-1]
        return self.write_excel(data, save_file_name)