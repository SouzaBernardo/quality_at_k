import openpyxl
class ExcelProcessor: 
    def __init__(self):
        pass



    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data
        except Exception as e:
            print(f"Error reading excel file: {e}")
            return None

    def write_excel(data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            for row in data:
                for cell in row:
                    ws.cell(row=cell[0], column=cell[1], value=cell[2])
            wb.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing to file: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        # Change the specified column in the Excel file to uppercase
        wb = openpyxl.load_workbook(save_file_name)
        ws = wb.active
        for i in range(N):
            row = ws.iter_rows(min_row=i+1, values_only=True)
            for cell in row:
                cell.value = cell.value.upper()
        wb.save(save_file_name)
        return 0, save_file_name