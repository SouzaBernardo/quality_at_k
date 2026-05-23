import openpyxl
class ExcelProcessor: 
    def __init__(self):
        pass



    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        # open the excel file
        wb = openpyxl.load_workbook(file_name)
        # get the active sheet
        sheet = wb.active
        # read the data
        data = []
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        # close the file
        wb.close()
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            # Get the active sheet
            ws = wb.active
            # Set the column names
            ws.append(data[0])
            # Iterate through the data
            for row in data[1:]:
                ws.append(row)
            # Save the workbook
            wb.save(file_name)
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        # open the source file
        wb = openpyxl.load_workbook(filename=save_file_name)
        # get the sheet
        sheet = wb.active
        # get the specified column
        column = sheet.columns[N-1]
        # change the column to uppercase
        for cell in column:
            cell.value = str(cell.value).upper()
        # save the file
        wb.save(save_file_name)
        # return the success and output file name
        return (1, save_file_name)