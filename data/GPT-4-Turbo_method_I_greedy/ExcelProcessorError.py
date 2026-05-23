import openpyxl
class ExcelProcessor: 
    def __init__(self):
        pass


    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data
    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"An error occurred: {e}")
            return 0
    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        # Read the data from the file
        data = self.read_excel(save_file_name)
    
        # Process the data
        for i in range(len(data)):
            row = list(data[i])
            row[N] = row[N].upper() if isinstance(row[N], str) else row[N]
            data[i] = tuple(row)
    
        # Write the processed data back to the file
        success = self.write_excel(data, save_file_name)
    
        return success, save_file_name