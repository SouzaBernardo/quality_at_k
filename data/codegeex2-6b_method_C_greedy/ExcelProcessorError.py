import openpyxl
class ExcelProcessor: 
    def __init__(self):
        pass



    def read_excel(self, file_name):
        

        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        data = []
        for row in ws.iter_rows():
            data.append([cell.value for cell in row])
        return data


    def write_excel(self, data, file_name):
        

        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(file_name)
        return 1


    def process_excel_data(self, N, save_file_name):
        

        pass