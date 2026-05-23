import openpyxl
class ExcelProcessor: 
    def __init__(self):
        pass



    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, max_row=N):
            data.append(row)
        return data


    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        try:
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
            ws.append(data)
            wb.save(file_name)
            return 1
        except Exception as e:
            print(e)
            return 0


    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        # read the excel file
        wb = openpyxl.load_workbook(save_file_name)
        ws = wb.active
        # get the column name
        column_name = ws.cell(row=1, column=N).value
        # change the column name to uppercase
        ws.cell(row=1, column=N).value = column_name.upper()
        # save the excel file
        wb.save(save_file_name)
        return 1, save_file_name